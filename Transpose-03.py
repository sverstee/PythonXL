#! python3
# transpose.py tranpose numbers
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

############
# Pandas and NumPy are for another test
############
import pandas as pd
import numpy as np

# Needed for Excel
import xlrd

########  COuldn't transpose this with PANDA
# dfs = pd.read_excel("c:\_scripts\python\produceSales1.xlsx", sheetname=None)
#####################

df = pd.read_excel(open('c:\_scripts\python\produceSales1.xlsx','rb'), 'Sheet')

print (df)

#########  Transpose
df1 = df.T
print (df1)

#TODO Write out content

wb2 = load_workbook('c:\_scripts\python\produceSales.xlsx')

# New to get rows
sheetName = wb2.get_sheet_by_name('Sheet')
rowCount = sheetName.max_row
columnCount = sheetName.max_column
print ('Sheet length Rows, Columns ', rowCount,' ', columnCount)

# Print our contents
print (wb2.get_sheet_names())

# Transpose

#Write Out
wb = Workbook()
dest_filename = 'c:\_scripts\python\output1.xlsx'
ws1 = wb.active

##############
# This seems inefficient - Need to be able to find range
# This is generic code
##############
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")

#TODO Fix this to work with get_HIghest_row
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)

# Save Code
wb.save(filename = dest_filename)
