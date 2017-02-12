#! python3

# Program to transpose rows and columms - SV Test

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
outWb = openpyxl.load_workbook('output.xlsx')
outSheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices


# Loop through the rows and update the prices.
for rowNum in range(1, sheet.get_highest_row()): # get contents
    produceName = sheet.cell(row=rowNum, column=1).value
    rowValue = sheet.cell(row=rowNum, column=1).value
    print(rowValue)
wb.save("updatedProduceSales.xlsx")
